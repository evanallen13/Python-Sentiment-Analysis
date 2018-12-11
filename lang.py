from openpyxl import *
from textblob import TextBlob

book1= load_workbook('Hotel_Reviews .xlsx',data_only=True)
sheet1= book1.active

for i in range(2,35913):
    #Review, Lang, Sent
        review= TextBlob(str(sheet1.cell(column=11,row=i).value))
        lang= TextBlob.detect_language(review)
        if lang != "en":
            review= review.translate(to="en")
            review= str(review)
            sheet1.cell(column=11,row=i).value= review
            
book1.save('Hotel_Reviews .xlsx')
            
    
