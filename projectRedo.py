from openpyxl import *
import csv 
from textblob import TextBlob
from projectClasses import RegSeason
from statistics import mean

westWinter=['West','Winter']
westSpring=['West','Spring']
westSummer=['West','Summer']
westFall=['West','Fall']

southwestWinter=['Southwest','Winter']
southwestSpring=['Southwest','Spring']
southwestSummer=['Southwest','Summer']
southwestFall=['Southwest','Fall']

southeastWinter=['Southeast','Winter']
southeastSpring=['Southeast','Spring']
southeastSummer=['Southeast','Summer']
southeastFall=['Southeast','Fall']

midwestWinter=['Midwest','Winter']
midwestSpring=['Midwest','Spring']
midwestSummer=['Midwest','Summer']
midwestFall=['Midwest','Fall']

northeastWinter=['Northeast','Winter']
northeastSpring=['Northeast','Spring']
northeastSummer=['Northeast','Summer']
northeastFall=['Northeast','Summer']

book1= load_workbook('Hotel_Reviews .xlsx',data_only=True)
sheet1= book1.active
monthDate= sheet1.cell(column=8,row=2).value
months=[]
sents=[]
for i in range(2,35913):#35913
    if sheet1.cell(column=11,row=i).value != None and sheet1.cell(column=8,row=i).value != None:
        monthDate= str(sheet1.cell(column=8,row=i).value)
        month= monthDate[5:7]
        months.append(month)
        if len(sheet1.cell(column=7,row=i).value)== 2:
            state= sheet1.cell(column=7,row=i).value
        else:
            state= sheet1.cell(column=13,row=i).value
        #Review, Lang, Sent
        review= TextBlob(str(sheet1.cell(column=11,row=i).value))
        sent= int(review.sentiment.polarity)
        sents.append(sent)
        reg=RegSeason.regions(state,i)#<<<<<<<<<<<<<<<<<<<<<<<<
        season= RegSeason.seasons(monthDate)#<<<<<<<<<<<<<<<<<<
        
        if reg == "West" and season== "Winter":
            westWinter.append(sent)
        elif reg == "West" and season== "Spring":
            westSpring.append(sent)
        elif reg == "West" and season== "Summer":
            westSummer.append(sent)
        elif reg == "West" and season== "Fall":
            westFall.append(sent)

        elif reg == "Midwest" and season== "Winter":
            midwestWinter.append(sent)
        elif reg == "Midwest" and season== "Spring":
            midwestSpring.append(sent)
        elif reg == "Midwest" and season== "Summer":
            midwestSummer.append(sent)
        elif reg == "Midwest" and season== "Fall":
            midwestFall.append(sent)

        elif reg == "Southeast" and season== "Winter":
            southeastWinter.append(sent)
        elif reg == "Southeast" and season== "Spring":
            southeastSpring.append(sent)
        elif reg == "Southeast" and season== "Summer":
            southeastSummer.append(sent)
        elif reg == "Southeast" and season== "Fall":
            southeastFall.append(sent)

        elif reg == "Southwest" and season== "Winter":
            southwestWinter.append(sent)
        elif reg == "Southwest" and season== "Spring":
            southwestSpring.append(sent)
        elif reg == "Southwest" and season== "Summer":
            southwestSummer.append(sent)
        elif reg == "Southwest" and season== "Fall":
            southwestFall.append(sent)
            
        elif reg == "Northeast" and season== "Winter":
            northeastWinter.append(sent)
        elif reg == "Northeast" and season== "Spring":
            northeastSpring.append(sent)
        elif reg == "Northeast" and season== "Summer":
            northeastSummer.append(sent)
        elif reg == "Northeast" and season== "Fall":
            northeastFall.append(sent)

#CSV
#open a csv file  and set it up for writing 
CSVoutfileW= open('Results.csv',"w")
CSVwriterRest= csv.writer(CSVoutfileW, delimiter=',',lineterminator='\n')

long= [westWinter,westSpring,westSummer,westFall,southwestWinter,southwestSpring,southwestSummer,southwestFall,southeastWinter,southeastSpring
    ,southeastSummer,southeastFall,midwestWinter,midwestSpring,midwestSummer,midwestFall,northeastWinter,northeastSpring,northeastSummer,northeastFall]

headers=['Region','Season','Reviews','Avg']
CSVwriterRest.writerow(headers)
for i in range(0,20):
    amt= len(long[i])
    kid= long[i]
    region= kid[0]
    season= kid[1]
    kid.pop(0)
    kid.pop(0)
    data= mean(kid)
    info= [region,season,amt,data]
    CSVwriterRest.writerow(info)

CSVoutfileW.close()

